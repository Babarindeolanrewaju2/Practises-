from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

# create a new Chrome browser instance
driver = webdriver.Chrome()

# navigate to the Zoopla website
driver.get("https://www.zoopla.co.uk/")

# wait for the page to load
time.sleep(5)

# enter a search query for properties in London
search_box = driver.find_element_by_name("q")
search_box.send_keys("London")
search_box.send_keys(Keys.RETURN)

# wait for the search results page to load
time.sleep(5)

# find all property listings on the page
listings = driver.find_elements_by_class_name("listing-results-wrapper")

# create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# write the scraped data to the worksheet
row_num = 1
for listing in listings:
    address = listing.find_element_by_class_name("listing-results-address").text
    price = listing.find_element_by_class_name("listing-results-price").text
    bedrooms = listing.find_element_by_class_name("num-icon.num-beds").text
    bathrooms = listing.find_element_by_class_name("num-icon.num-baths").text

    worksheet.cell(row=row_num, column=1, value=address)
    worksheet.cell(row=row_num, column=2, value=price)
    worksheet.cell(row=row_num, column=3, value=bedrooms)
    worksheet.cell(row=row_num, column=4, value=bathrooms)

    row_num += 1

# save the workbook to a file
workbook.save("zoopla_properties.xlsx")

# close the browser
driver.close()
